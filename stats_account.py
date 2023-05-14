import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pymongo import MongoClient
import pandas as pd
from datetime import datetime, timedelta
import re
from bson.objectid import ObjectId
from decimal import Decimal
import calendar
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from xlsxwriter import Workbook
import os
from openpyxl import load_workbook
import math
import google.auth
# from GoogleApiSupport.auth import get_service
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from apiclient import discovery
from httplib2 import Http
from oauth2client import file, client, tools
import config

client = MongoClient(f'mongodb+srv://{config.mongo_pat}')
db = client['legacy-api-management']
col_soc = db["societies"]
col_it = db["items"]
col_users = db["users"]



def get_card(ent):
    l_id, l_member, l_name = [], [], []
    search = ent
    search_expr = re.compile(f".*{search}.*", re.I)
    cursor = col_soc.find({'name': {'$regex': search_expr}})
    for rep in cursor:
        l_member.append(len(rep["members"]))
        l_id.append(rep["_id"])
        l_name.append(rep["name"])

    df = pd.DataFrame(list(zip(l_id, l_member, l_name)), columns=['id', 'members', 'name'])
    df = (df[df.members == df.members.max()])
    my_id = (df['id'].values[0])

    l_username = []
    cursor_soc = col_soc.find({'_id': my_id})

    l_card = []
    nb_card = []
    for c in cursor_soc:
        len_members = (len(c["members"]))
        l_members = [c['members'][x]['user'] for x in range(len_members)]
        for i in l_members:
            cursor_user = col_users.find({'_id': i})
            for d in cursor_user:
                len_sub = (len(d['travelSubscriptions']))
                l_username.append(d['username'])
                l_res = []
                nb_card.append(len_sub)
                if len_sub > 0:
                    for ls in range(len_sub):
                        try:label = d['travelSubscriptions'][ls]['label']
                        except:label = ""
                        try:
                            endAt = d['travelSubscriptions'][ls]['endAt']
                            endAt2 = endAt.strftime("%Y-%m-%d %H:%M")
                        except: endAt2 = ""
                        try :type = d['travelSubscriptions'][ls]['type']
                        except: type =""
                        res = label,endAt2,type
                        l_res.append(res)
                else:
                    l_res.append("-")
                l_card.append(l_res)
    print(len(l_card),len(l_username),len(l_members))
    org_df = pd.DataFrame(list(zip(l_members,l_username,nb_card,l_card)), columns=['user','user_id',"nb_card","card/EndAt/type"])
    org_df.to_excel(f"csv/{ent}_cards.xlsx")

def no_col_itemid(filename):
    df = pd.read_csv(f'csv/BR/br_in_queue/{filename}.csv',encoding='UTF-8',delimiter=";")
    df = df.fillna(0)
    df['TYPE'] = df['TYPE'].replace("RentalCar","Car")

    l_itemid = []
    l_provider = []
    l_type = []
    l_user = []

    for i in range(len(df)):
        status = df['STATUS'][i]
        type = df['TYPE'][i].lower()
        created = df['BOOKING_CREATED_AT'][i]
        createdAt = datetime.strptime(created, '%d/%m/%Y')
        createdAt_end = createdAt + timedelta(days=1)
        item_id = ""
        provider = ""
        userid = df['USER_ID'][i]

        try:
            datein = df['CHECKIN'][i]
            datein = datetime.strptime(datein, '%d/%m/%Y')
            datein = (str(datein)[:10])
            dateout = df['CHECKOUT'][i]
            dateout = datetime.strptime(dateout, '%d/%m/%Y')
            dateout = (str(dateout)[:10])
        except:
            dateout = ""
            datein = ""

        if userid != 0 :
            if type == "hotel" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.datein.date": datein,
                        # "detail.dateout.date": dateout,
                        # "status" : status
                    })
                    for c in cursor:
                        item_id = (c['id'])
                        provider = (c['provider'])
                        print(userid,item_id)
            elif type == "flight" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.datein.date": datein,
                        # "detail.dateout.date": dateout,
                        # "status" : status
                    })
                    for c in cursor:
                        item_id = (c['id'])
                        provider = (c['provider'])
            elif type == "train" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.journeys.departure.date.date": datein,
                        # "detail.journeys.arrival.date.date": dateout
                    })
                    for c in cursor:
                        item_id = (c['id'])
                        provider = (c['provider'])
            elif type == "car" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.pickupDate.date": datein,
                        # "detail.returnDate.date": dateout
                    })
                    for c in cursor:
                        item_id = (c['id'])
                        provider = (c['provider'])
            elif type == "transfer" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end}
                    })
                    for c in cursor:
                        item_id = (c['id'])
                        provider = (c['provider'])
            if provider == "offline":
                provider = "true"
            else:
                provider = "false"
        else :
            item_id= ""
            provider= ""

        l_itemid.append(item_id)
        l_provider.append(provider)
        l_type.append(type)
        l_user.append(userid)
    df['ITEM_ID'] = l_itemid
    df['USERID'] = l_user
    df['IS_OFFLINE'] = l_provider
    df['TYPE'] = l_type
    df.to_csv(f'csv/BR/br_in_queue/{filename}_v2.csv')

def col_itemid_missrow(filename):
    df = pd.read_csv(f'csv/BR/br_in_queue/{filename}.csv',encoding='UTF-8',delimiter=";")
    df = df.fillna(0)
    df['TYPE'] = df['TYPE'].replace("RentalCar","Car")
    df_missed = df.loc[df['ITEM_ID'] == 0].reset_index()
    df_ok = df.loc[df['ITEM_ID'] != 0].reset_index()
    df_missed.to_csv(f'csv/BR/br_in_queue/{filename}_missed.csv')

    l_itemid = []
    l_type = []
    l_user = []

    for i in range(len(df_missed)):
        userid = df['USERID'][i]
        user_id = df['USER_ID'][i]
        name = df['USER'][i]
        if userid != 0 and user_id == 0:
            userid = userid
        elif userid == 0 and user_id != 0:
            userid = df['USER_ID'][i]
        elif userid == 0 and user_id == 0:
            userid = df.loc[(df['USER'] == name) & (df['USERID'] != 0), "USERID"].values[0]
            print(userid)
        type = df['TYPE'][i].lower()
        created = df['BOOKING_CREATED_AT'][i]
        createdAt = datetime.strptime(created, '%d/%m/%Y')
        createdAt_end = createdAt + timedelta(days=1)
        item_id = ""

        try:
            datein = df['CHECKIN'][i]
            datein = datetime.strptime(datein, '%d/%m/%Y')
            datein = (str(datein)[:10])
            dateout = df['CHECKOUT'][i]
            dateout = datetime.strptime(dateout, '%d/%m/%Y')
            dateout = (str(dateout)[:10])
        except:
            dateout = ""
            datein = ""

        if userid != 0 :
            if type == "hotel" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.datein.date": datein,
                        # "detail.dateout.date": dateout,
                    })
                    for c in cursor:
                        item_id = (c['id'])
                        print(userid,item_id)
            elif type == "flight" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.datein.date": datein,
                        # "detail.dateout.date": dateout,
                    })
                    for c in cursor:
                        item_id = (c['id'])
            elif type == "train" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.journeys.departure.date.date": datein,
                        # "detail.journeys.arrival.date.date": dateout
                    })
                    for c in cursor:
                        item_id = (c['id'])
            elif type == "car" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end},
                        # "detail.pickupDate.date": datein,
                        # "detail.returnDate.date": dateout
                    })
                    for c in cursor:
                        item_id = (c['id'])
            elif type == "transfer" :
                    cursor = col_it.find({
                        'type': type,
                        'travelers.userId': ObjectId(userid),
                        "createdAt": {"$gte": createdAt, "$lt": createdAt_end}
                    })
                    for c in cursor:
                        item_id = (c['id'])
        else :
            item_id= ""

        l_itemid.append(item_id)
        l_type.append(type)
        l_user.append(userid)
    df_missed['ITEM_ID'] = l_itemid
    df_missed['USERID'] = l_user
    df_missed['TYPE'] = l_type

    liste = [df_missed, df_ok]
    frame = pd.concat(liste, axis=0, join="inner")
    frame = frame.reset_index()
    frame = frame.drop(columns =["level_0",'index'])
    frame.to_csv(f'csv/BR/br_in_queue/{filename}_v3.csv')

def br2(filename):
    try: os.mkdir(f'csv/BR/br_result/{filename}')
    except: pass
    try: os.mkdir(f'csv/del/{filename}')
    except: pass

    df = pd.read_csv(f'csv/BR/br_in_queue/{filename}.csv',encoding='UTF-8',delimiter=";")
    df = df.fillna(0)
    print("df ok")
    ### Convert str in float
    l_billed = []
    l_factured = []
    l_new_billed = []
    l_itemid = []
    l_type = []
    l_anticip = []
    l_status= []
    for i in range(len(df)):
        try:
            billed = df['TOTAL_BILLED'][i]
        except:
            billed = 0
        try:
            factured = df['FACTURED'][i]
        except:
            factured = 0
        try:
            billed = billed.replace(",", '.')
            billed = int(billed)
        except:
            billed = float(billed)
        if billed > 0:
            billed = billed
        elif billed < 0:
            billed = billed
        elif billed == 0:
            billed = 0

        try:
            factured = factured.replace(",", '.')
            factured = int(factured)
        except:
            factured = float(factured)
        if factured > 0:
            factured = factured
        elif billed < 0:
            factured = factured
        elif factured == 0:
            factured = 0

        l_billed.append(billed)
        l_factured.append(factured)
        ### New colonne for merge "billed" & "factured" if billed 0
        if factured > 0:
            billed = factured
        else:
            billed = billed

        l_new_billed.append(billed)
        status = df['STATUS'][i]
        l_status.append(status)
        KO = ['cancelled','modified',"CANCELED","MODIFIED"]
        if status in KO:
            l_anticip.append(0)
        else:
            if df['ANTICIPATION_IN_DAYS'][i] >0:
                l_anticip.append(df['ANTICIPATION_IN_DAYS'][i])
            else:
                l_anticip.append(0)
        itemid = df['ITEM_ID'][i]
        if df['ITEM_ID'][i] == 0 and df['BOOKING_ID'][i]:
            itemid = (df['BOOKING_ID'][i])
        l_itemid.append(itemid)
    ### split other
        type = df['TYPE'][i]
        if type == 'other':
            if "Visa Connect" in df['PLACE'][i]:
                type = "other_visa"
            else:
                type = "other_card"
        l_type.append(type)
    #### append columns
    df['ITEM_ID'] = l_itemid
    df['NEW_TYPE'] = l_type
    df['ANTICIPATION_IN_DAYS'] = l_anticip
    df['STATUS'] = l_status
    df['TOTAL_BILLED'] = l_billed
    df['FACTURED'] = l_factured
    df['NEW_TOTAL_BILLED'] = l_new_billed

    # #################
    df['NEW_TYPE'] = df['NEW_TYPE'].str.replace('rentalcar', 'car')

    ### keep only item_id not null
    df_with_idtems = df.loc[df['ITEM_ID'] !=0 & (df['USERID'] != 0)].reset_index()
    df_with_idtems.to_csv(f"csv/del/{filename}/out_{filename}_1x1.csv")

    ### keep only item_id not null
    df_noidtems_nouser = df.loc[(df['ITEM_ID'] == 0) & (df['USERID'] == 0)].reset_index()
    df_noidtems_nouser.to_csv(f"csv/del/{filename}/out_{filename}_0x0.csv")

    ### get only item_id null

    df_no_idtems = df.loc[(df['ITEM_ID'] == 0) & (df['USERID'] != 0)].reset_index()
    df_no_idtems.to_csv(f"csv/del/{filename}/out_{filename}_0x1.csv")

    print("search")

    l_new_item = []
    for g in range(len(df_no_idtems)):
        pnr = (df_no_idtems['PLACE'][g]).split(" - ")[0].strip()
        type = (df_no_idtems['NEW_TYPE'][g])
        createdAt = (df_no_idtems['BOOKING_CREATED_AT'][g])
        createdAt = datetime.strptime(createdAt, '%d/%m/%Y')
        createdAt_end = createdAt + timedelta(days=1)
        userid = (df_no_idtems['USERID'][g])

        if type == "train" and len(pnr) == 6:

            try:
                cursor = col_it.find(
                {'type': type, 'detail.id': pnr, "createdAt": {"$gte": createdAt, "$lt": createdAt_end}})
                for c in cursor:
                    id_item = c['id']
                    print('train', id_item)
            except:
                id_item = "!!!"
        elif type == "hotel" and len(pnr) == 13:
            cursor = col_it.find({'type': type, 'detail.supplierbookingref': pnr,
                                  "createdAt": {"$gte": createdAt, "$lt": createdAt_end}})
            for c in cursor:
                id_item = c['id']
                print("hotel", id_item)
        elif type == "flight" and userid != "0":
            cursor = col_it.find({'type': type, 'travelers.userId': ObjectId(userid),
                                  "createdAt": {"$gte": createdAt, "$lt": createdAt_end}})
            for c in cursor:
                id_item = c['id']
        else:
            id_item = "!!!"
        l_new_item.append(id_item)
    print("search finito")
    df_no_idtems['ITEM_ID'] = l_new_item

    df_no_idtems.to_csv(f"csv/del/{filename}/out_{filename}_0x1.csv")
    print("Ok")

    df_no_idtems_nores = df_no_idtems.loc[(df_no_idtems['ITEM_ID'] == '!!!')].reset_index()
    df_no_idtems_nores.to_csv(f"csv/del/{filename}/out_{filename}_!!!x0.csv")

    df_no_idtems = df_no_idtems.loc[(df_no_idtems['ITEM_ID'] != '!!!')].reset_index()
    df_no_idtems.to_csv(f"csv/del/{filename}/out_{filename}_0x1.csv")

    liste = [df_with_idtems,df_no_idtems]
    frame = pd.concat(liste, axis=0, join="inner")
    frame = frame.reset_index()
    frame.to_csv(f"csv/BR/br_result/{filename}/id_ok.csv")

    liste2 = [df_noidtems_nouser,df_no_idtems_nores]
    frame = pd.concat(liste2, axis=0, join="inner")
    frame = frame.reset_index()
    frame.to_csv(f"csv/BR/br_result/{filename}/id_ko.csv")

def br_group(filename):
    df = pd.read_csv(f"csv/BR/br_result/{filename}/id_ok.csv")
    df_code = pd.read_csv(f'const/AO&D _2023.csv')

    nbll, nadv = [],[]
    for i in range (len(df)):
        billed = df['NEW_TOTAL_BILLED'][i]
        try:
            billed = billed.replace(",", '.')
            billed = int(billed)
        except:
            billed = np.around(float(billed),2)
        nbll.append(billed)
        adv = df['ANTICIPATION_IN_DAYS'][i]
        try:
            adv = adv.replace(",", '.')
            adv = int(adv)
        except:
            adv = np.around(float(adv),0)
        nadv.append(adv)
    df['ANTICIPATION_IN_DAYS']=nadv
    df['NEW_TOTAL_BILLED']=nbll

    l_nb_legs,l_ori, l_des, l_trajet, l_od, l_four = [], [], [], [], [], []
    l_class,l_anti = [], []
    l_trav,l_delta_days,l_nb_trav = [],[],[]
    l_count = []
    for i in range(len(df)):
        l_pdt= []
        l_travelClass = []
        item_id = df['ITEM_ID'][i]
        type = (df['NEW_TYPE'][i]).lower()

        cursor = col_it.find({'id': item_id})
        if item_id !="" or item_id != 0:
            for j in cursor:
                if type == 'hotel':
                    nbre = 1
                    try:
                        travelClass = (j['detail']['hotelRating'])
                    except:
                        travelClass = "nc"
                    four = (j['detail']['hotelTitle']).upper()
                    try:
                        pdt = (j['detail']['hotelCity']).upper()
                    except:
                        pdt = (j['detail']['hotelAddress'])
                    l_pdt.append(pdt)
                    l_travelClass.append(travelClass)

                    ori = ""
                    des = pdt
                elif type == 'train':
                    if j["type"] == "other":
                        l_pdt.append(j["detail"]["title"])
                        l_travelClass.append("")
                        nbre = 1
                        four = ''
                    else:
                        nbre = (len(j['detail']['journeys']))
                        for l in range(nbre):
                            four = (j['detail']['journeys'][0]['segments'][0]['transport']['trainCode'])
                            try:
                                ori = j['detail']['journeys'][l]['departure']['city'].upper()
                                des = j['detail']['journeys'][l]['arrival']['city'].upper()

                                if ori == "" or des == "":
                                    ori = j['detail']['journeys'][l]['departure']['name'].upper()
                                    des = j['detail']['journeys'][l]['arrival']['name'].upper()
                            except:
                                ori = j['detail']['journeys'][l]['departure']['name'].upper()
                                des = j['detail']['journeys'][l]['arrival']['name'].upper()
                            try:
                                travelClass = j['detail']['journeys'][l]['travelClass']
                            except:
                                travelClass = "nc"
                            city = ["PARIS","LYON","MARSEILLE","BORDEAUX"]
                            for c in city:
                                if c in ori:
                                    ori = c
                                elif c in des:
                                    des = c
                            ori_des = ori, des
                            l_pdt.append(ori_des)
                            l_travelClass.append(travelClass)
                elif type == 'flight':
                    print(item_id)

                    nbre = (len(j['detail']['trips'][0]['legs']))

                    for l in range(nbre):
                        ori0 = j['detail']['trips'][0]['legs'][l]['departure']['city']
                        des0 = j['detail']['trips'][0]['legs'][l]['arrival']['city']
                        four = (j['detail']['trips'][0]['legs'][l]['governingCarrier'])

                        l_cabin = [("Premium First", "P"), ("First", "F"), ("Premium Business", "J"), ("Business", "C"),
                                   ("Premium Economy", "S"), ("Economy", "Y")]
                        try:
                            travelClass = j['detail']['trips'][0]['legs'][l]['cabinCodes']
                            for c, p in l_cabin:
                                if p in travelClass:
                                    travelClass = c
                        except:
                            try:
                                nbr_s = len(j['detail']['trips'][0]['legs'][l]["schedules"])
                                for ns in range(nbr_s):
                                    travelClass = j['detail']['trips'][0]['legs'][l]["schedules"][ns]['cabinCode']
                                    for c, p in l_cabin:
                                        if p in travelClass:
                                            travelClass = c
                            except:
                                travelClass = "nc"

                        try:
                            ori = df_code.loc[df_code['Origin Code'] == ori0, 'Label cities of origin'].values[0].upper()
                        except:
                            try:
                                ori = df_code.loc[df_code['Destination Code'] == ori0, 'Label cities of destination'].values[0].upper()
                            except:
                                ori =("!!!")
                        try:
                            des = df_code.loc[df_code['Destination Code'] == des0, 'Label cities of destination'].values[0].upper()
                        except:
                            try:
                                des = df_code.loc[df_code['Origin Code'] == des0, 'Label cities of origin'].values[0].upper()
                            except:
                                des =("!!!")
                        ori_des = (ori, des)
                        l_pdt.append(ori_des)
                        l_travelClass.append(travelClass)
                elif type == 'car':
                    type = "car"
                    nbre = 1
                    l_travelClass.append("")
                    try:
                        four = j['formData']['agencyName']
                    except:
                        four = "nc"
                    ori = j['detail']['pickupAddress']['city']
                    des = j['detail']['returnAddress']['city']
                    pdt = (ori, des)
                    l_pdt.append(pdt)
                elif type == 'fee':
                    l_travelClass.append("")
                    nbre = 1
                    four = ""
                    ori = ""
                    des = ""
                    ori_des = ("", "")
                    l_pdt.append(ori_des)
                elif type == 'transfer':
                    l_travelClass.append("")
                    nbre = 1
                    four = ""
                    ori = ""
                    des = ""
                    ori_des = ("", "")
                    l_pdt.append(ori_des)
                else:
                    nbre =1
                    ori = ""
                    des = ''
                    ori_des = ("", "")

                    four = ""
                    l_pdt.append(ori_des)

                if (nbre == 2 or nbre == 1) and (type == "train" or type == "flight"):
                    ori = l_pdt[0][0]
                    des = l_pdt[0][1]
                elif nbre == 3 or nbre == 4 and (type == "train" or type == "flight"):
                    ori = l_pdt[0][0]
                    des = l_pdt[1][1]

        if item_id == 0 or item_id == "":
            if type == "hotel" or type =="car" or type =="fee" or type =="transfer":
                nbre = 1
            elif type == "flight":
                try:
                    nbre =(len(df['PLACE'][i].split("->")))
                except:
                    nbre =(len(df['PLACE'][i].split("=>")))
            elif type == "train":
                try:
                    nbre =(len(df['PLACE'][i].split("->")))
                except:
                    try:
                        nbre =(len(df['PLACE'][i].split("=>")))
                    except:
                        nbre = (len(df['PLACE'][i].split(" - ")))

            ori = ""
            des = ''
            ori_des = ("", "")
            four = "nc"
            l_pdt.append(ori_des)

        checkin = df['CHECKIN'][i]
        try:
            checkin = datetime.strptime(checkin,'%d/%m/%Y')
        except:
            checkin = ""
        checkout = df['CHECKOUT'][i]
        try:
            checkout = datetime.strptime(checkout, '%d/%m/%Y')
        except:
            checkout = ""
        try:
            delta_days = ((checkout-checkin).days)
        except:
            delta_days = ""

        status = df['STATUS'][i]
        KO = ['cancelled', 'modified', "CANCELED", "MODIFIED"]
        if status in KO:
            nbre = -nbre
            l_count.append(0)
        else:
            nbre = nbre
            l_count.append(1)
        l_four.append(four)
        l_trajet.append(l_pdt)
        l_nb_legs.append(nbre)
        l_ori.append(ori)
        l_des.append(des)
        l_od.append(str(f'{ori} {des}').strip())
        l_class.append(str(l_travelClass).replace('[','').replace(']','').replace("'",''))
        l_delta_days.append(delta_days)

    df['COUNT'] = l_count
    df['NB_LEGS'] = l_nb_legs
    df['LEGS'] = l_trajet
    df['ORI'] = l_ori
    df['DES'] = l_des
    df['O&D'] = l_od
    df['CLASS'] = l_class
    df["FOURNISSEUR"] = l_four
    df['DUREE'] = l_delta_days

    df.to_csv(f"csv/BR/br_result/{filename}/id_ok.csv")

def off_on(filename):
    df = pd.read_csv(f"csv/BR/br_result/{filename}/id_ok.csv")
    l_leg_off, l_ca_off,l_leg_on, l_ca_on = [],[],[],[]
    for i in range (len(df)):
        is_off = df['IS_OFFLINE'][i]
        print(type(is_off))
        nb_leg = df['NB_LEGS'][i]
        billed = df['NEW_TOTAL_BILLED'][i]
        if is_off == True or is_off == "true" or is_off == "True":
            l_leg_off.append(nb_leg)
            l_ca_off.append(billed)
            l_leg_on.append(0)
            l_ca_on.append(0)
        elif is_off == False or is_off == "false" or is_off == "False":
            l_leg_off.append(0)
            l_ca_off.append(0)
            l_leg_on.append(nb_leg)
            l_ca_on.append(billed)
        else:
            l_leg_off.append("")
            l_ca_off.append("")
            l_leg_on.append("")
            l_ca_on.append("")
    df['CA_offline'] = l_ca_off
    df['LEGS_offline'] = l_leg_off
    df['CA_online'] = l_ca_on
    df['LEGS_online'] = l_leg_on
    df = df.drop(columns=['Unnamed: 0.1','Unnamed: 0','level_0','index'])
    df.to_csv(f"csv/BR/br_result/{filename}/id_ok.csv")

def top(filename):
    df = pd.read_csv(f"csv/BR/br_result/{filename}/id_ok.csv")
    #Resume total depenses & avg
    my_type = ['train','car','flight','hotel','transfer']
    l_type = []
    l_sum_billed = []
    l_avg_anticip = []
    l_total_ca_off = []
    l_total_ca_on = []
    l_total_legs_off = []
    l_total_legs_on = []
    l_total_legs = []

    for mt in my_type:
        df_type = df.loc[df['NEW_TYPE'] == mt]
        sum_billed = (df_type.sum(numeric_only=True)["NEW_TOTAL_BILLED"]).round(0)
        avg_anticip = df_type.mean(numeric_only=True)["ANTICIPATION_IN_DAYS"].round(0)
        sum_ca_off = (df_type.sum(numeric_only=True)["CA_offline"]).round(0)
        sum_ca_on = (df_type.sum(numeric_only=True)["CA_online"]).round(0)
        sum_legs_off = (df_type.sum(numeric_only=True)["LEGS_offline"]).round(0)
        sum_legs_on = (df_type.sum(numeric_only=True)["LEGS_online"]).round(0)
        sum_legs = (df_type.sum(numeric_only=True)["NB_LEGS"]).round(0)
        l_type.append(mt)
        l_sum_billed.append(sum_billed)
        l_total_legs.append(sum_legs)
        l_total_legs_off.append(sum_legs_off)
        l_total_legs_on.append(sum_legs_on)
        l_total_ca_on.append(sum_ca_on)
        l_total_ca_off.append(sum_ca_off)
        l_avg_anticip.append(avg_anticip)
        print(sum_billed)

        # ONLINE/OFFLINE

    df_calc = pd.DataFrame(list(zip(l_type, l_sum_billed, l_total_legs, l_total_legs_off, l_total_legs_on,l_total_ca_off,l_total_ca_on,l_avg_anticip)),
                            columns=['Type', 'total_depenses', "nbre_trajet", "nbre_trajet_offline", "nbre_trajet_online", "ca_offline", "ca_online", 'avg anticipation'])

    df_calc["avg_dep"]= (df_calc['total_depenses']/df_calc['nbre_trajet']).round(2)
    df_calc.loc['total'] = df_calc.sum(numeric_only=True)
    df_calc.to_csv(f"csv/BR/br_result/{filename}/calc_resa.csv")

    ###
    other_type = ['fee','other_card','other_visa']
    l_type = []
    l_sum_billed = []
    l_total_legs = []
    for mt in other_type:
        df_type = df.loc[df['NEW_TYPE'] == mt]
        sum_billed = (df_type.sum(numeric_only=True)["NEW_TOTAL_BILLED"]).round(0)
        sum_legs = (df_type.sum(numeric_only=True)["NB_LEGS"]).round(0)
        l_type.append(mt)
        l_sum_billed.append(sum_billed)
        l_total_legs.append(sum_legs)

    df_calc_other = pd.DataFrame(list(zip(l_type, l_sum_billed, l_total_legs)),
                            columns=['Type', 'total_depenses', "nbre"])
    df_calc_other.loc['total'] = df_calc_other.sum(numeric_only=True)
    df_calc_other["avg_dep"]= (df_calc_other['total_depenses']/df_calc_other['nbre']).round(2)

    df_calc_other.to_csv(f"csv/BR/br_result/{filename}/calc_other.csv")

    top_user = df.groupby(['USER','USERID']).sum(numeric_only=True).reset_index()

    top_user_only = (top_user[['USER','USERID','NEW_TOTAL_BILLED','COUNT']]).sort_values(by='NEW_TOTAL_BILLED',ascending=False).reset_index()

    for mt in my_type:
        df_type = df.loc[df['NEW_TYPE'] == mt]
        top_user_type = df_type.groupby(['USER', 'USERID']).sum(numeric_only=True).sort_values(by='NEW_TOTAL_BILLED',ascending=False).reset_index()
        top_user_type = top_user_type[['USER', 'USERID', 'NEW_TOTAL_BILLED','COUNT']].sort_values(by='NEW_TOTAL_BILLED',ascending=False)
        top_user_type.to_csv(f'csv/BR/br_result/{filename}/top_user_{mt}.csv')


    # # # GET ABO. CARD
    l_len_sub = []
    l_info_card = []

    for i in range(len(top_user_only)):
        user_name = top_user_only['USER'][i]
        cursor_user = col_users.find({'username': user_name})
        l_card = []
        for c in cursor_user:
            if user_name != 0:
                try:
                    len_sub = (len(c['travelSubscriptions']))
                    l_len_sub.append(len_sub)
                    for ls in range(len_sub):
                        type = (c['travelSubscriptions'][ls]['type'])
                        endAt = (c['travelSubscriptions'][ls]['endAt'])
                        label = (c['travelSubscriptions'][ls]['label'])
                        l_card.append(f'{type},{endAt},{label}')
                except:
                    l_card.append('')
            else:
                try:
                    user_id = top_user_only['USERID'][i]
                    cursor_user = col_users.find({'_id': ObjectId(user_id)})
                    for c in cursor_user:
                        try:
                            len_sub = (len(c['travelSubscriptions']))
                            l_len_sub.append(len_sub)
                            for ls in range(0,len_sub):
                                type = (c['travelSubscriptions'][ls]['type'])
                                endAt = (c['travelSubscriptions'][ls]['endAt'])
                                label = (c['travelSubscriptions'][ls]['label'])
                                l_card.append(f'[{label} > {type} > End : {endAt}]')
                        except:
                            l_card.append('')
                except:
                    l_card.append('')
        l_info_card.append(l_card)
    top_user_only["info_card"] = l_info_card
    top_user_only = top_user_only[['USER','USERID','NEW_TOTAL_BILLED','COUNT','info_card']]
    top_user_only.to_csv(f'csv/BR/br_result/{filename}/top_user.csv')

def calc(filename):
    df = pd.read_csv(f"csv/BR/br_result/{filename}/id_ok.csv")
    # df = df.loc[df['count'] != 0].reset_index()
    l_zero, l_first, l_second, l_third, l_fourth, l_fifth, = [],[],[],[],[],[]
    for c in range (len(df)):
            cl = (df["CLASS"][c])
            ty = df['NEW_TYPE'][c]
            if ty == "train":
                try:
                    first = cl.count('FIRST_CLASS')
                    second = cl.count('SECOND_CLASS')
                    third = 0
                    fourth = 0
                    fifth = 0
                    zero = 0
                except:
                    first = 0
                    second = 0
                    third = 0
                    fourth = 0
                    fifth = 0
                    zero = 0
            elif ty == "flight":
                try:
                    first = cl.count('Premium First')
                    second = cl.count('First')
                    third = cl.count('Premium Business')
                    fourth = cl.count('Business')
                    fifth = cl.count('Premium Economy')
                    zero = cl.count('Economy')
                except:
                    first = 0
                    second = 0
                    third = 0
                    fourth = 0
                    fifth = 0
                    zero = 0
            elif ty == "hotel":
                try:
                    first = cl.count('5')
                    second = cl.count('4')
                    third = cl.count('3')
                    fourth = cl.count('2')
                    fifth = cl.count('1')
                    zero = cl.count('0')
                except:
                    first = 0
                    second = 0
                    third = 0
                    fourth = 0
                    fifth = 0
                    zero = 0
            elif ty == "fee" or ty == "other_card" or ty == "other_visa" or ty== "car":
                first = 0
                second = 0
                third = 0
                fourth = 0
                fifth = 0
                zero = 0
            else:
                first = 0
                second = 0
                third = 0
                fourth = 0
                fifth = 0
                zero = 0
            l_first.append(first)
            l_second.append(second)
            l_third.append(third)
            l_fourth.append(fourth)
            l_fifth.append(fifth)
            l_zero.append(zero)
    df['zero'] = l_zero
    df['first'] = l_first
    df['second'] = l_second
    df['third'] = l_third
    df['fourth'] = l_fourth
    df['fifth'] = l_fifth

    # df.to_csv(f"csv/BR/br_result/{filename}/count_class.csv")

    # GET ID_SOC, MEMBERS, SUB_PRICE
    for i in range(1):
        item_id = df['ITEM_ID'][i]
        cursor = col_it.find({'id': item_id})
        for c in cursor:
            id_soc = (c['society']['_id'])
    cursor_soc = col_soc.find({'_id': id_soc})
    for c in cursor_soc:
        members = (c['members'])
        sub_price = (c['sub_price'])
        price_fee = (c['priceNormalFeeDefault'])
        createdAt = (str(c['createdAt'])[:10])
        firstbillfee = (str(c['createdAt'])[:10])
    nb_members = (len(members))
    df_resume = pd.DataFrame({'createdAt': [createdAt],'nb_members':[nb_members],'abo_price':[sub_price],'price_fee':[price_fee],'firstbillfee':[firstbillfee]},
                          index=[0])

    df_resume.to_csv(f"csv/BR/br_result/{filename}/resume.csv")

    ## GET LEGS & BILLED BY CLASS &  BY TYPE
    list_class = ['zero', 'first', 'second', 'third', 'fourth', 'fifth']
    name_class = [("train zero","nc"),
("train first","FIRST_CLASS"),
("train second","SECOND_CLASS"),
("hotel zero","0*"),
("hotel first","1*"),
("hotel second","2*"),
("hotel third","3*"),
("hotel fourth","4*"),
("hotel fifth","5*"),
("flight zero","Econony"),
("flight first","Premium Economy"),
("flight second","Business"),
("flight third","Premium Business"),
("flight fourth","First"),
("flight fifth","Premium First")]

    df_hotel = df.loc[(df['NEW_TYPE']=='hotel') & (df['NEW_TYPE']=='hotel')].sum(numeric_only=True).round(2)
    df_train = df.loc[df['NEW_TYPE']=='train'].sum(numeric_only=True).round(2)
    df_flight = df.loc[df['NEW_TYPE']=='flight'].sum(numeric_only=True).round(2)
    l_type = [("train",df_train),("hotel",df_hotel),("flight",df_flight)]
    l_nb_resa,l_t_resa, l_ca_resa, l_ty_resa= [],[],[],[]
    for t,v in l_type:
        for i in list_class:
            v = df.loc[(df['NEW_TYPE']==t) & (df[i])!=0].sum(numeric_only=True).round(2)
            ty_resa = (f'{t} {i}')
            for n,m in name_class:
                if ty_resa == n:
                    ty_resa = m
            if v['NB_LEGS'] != 0:
                nb_resa = v['NB_LEGS']
                ca_resa = v['NEW_TOTAL_BILLED']
                print(t, ty_resa, nb_resa, ca_resa)
                l_ty_resa.append(ty_resa)
                l_t_resa.append(t)
                l_nb_resa.append(nb_resa)
                l_ca_resa.append(ca_resa)
            else:
                pass
    df_class = pd.DataFrame(list(zip(l_t_resa,l_ty_resa,l_nb_resa,l_ca_resa)), columns=['Type','Classe',"NB_RESA",'CA'])
    df_class.to_csv(f"csv/BR/br_result/{filename}/top_classes.csv")

    # # TOP CARRIER
    df_four = df.groupby(["NEW_TYPE",'FOURNISSEUR'])[["NEW_TOTAL_BILLED","NB_LEGS"]].sum(numeric_only=True).round(1).reset_index()
    #
    df_car = pd.DataFrame()
    df_train = pd.DataFrame()
    df_hotel = pd.DataFrame()
    df_flight = pd.DataFrame()
    l_type = [(df_car,"car"), (df_train,'train'), (df_hotel,'hotel'), (df_flight,'flight')]
    l_top =[]

    for a,l in l_type:
        a = df_four.loc[df_four['NEW_TYPE'] == l].sort_values(by='NEW_TOTAL_BILLED',ascending=False).reset_index()
        # print(a)
        a = a.drop(columns="index")
        a =(a[:50])
        miss_row = 50 - (len(a))
        for i in range(miss_row):
            a.loc[len(a)] = ["","","",""]
            # print(a)
        a.loc[len(a)+1] = ["*****", f"END TOP FOURNI. {l.upper()}","*****" , "*****"]
        a.loc[len(a) + 2] = ["", "", "", ""]

        l_top.append(a)

    frame = pd.concat(l_top, axis=0, join="inner")
    frame.to_csv(f"csv/BR/br_result/{filename}/top_fournisseur.csv")

       # # TOP city
    df_four = df.groupby(["NEW_TYPE", 'O&D'])[["NEW_TOTAL_BILLED", "NB_LEGS"]].sum(numeric_only=True).round(1).reset_index()
    #
    df_car = pd.DataFrame()
    df_train = pd.DataFrame()
    df_hotel = pd.DataFrame()
    df_flight = pd.DataFrame()
    l_type = [(df_car, "car"), (df_train, 'train'), (df_hotel, 'hotel'), (df_flight, 'flight')]
    l_top = []

    for a, l in l_type:
        a = df_four.loc[df_four['NEW_TYPE'] == l].sort_values(by='NEW_TOTAL_BILLED', ascending=False).reset_index()
        a = a.drop(columns="index")
        a = (a[:10])
        miss_row = 10 - (len(a))
        for i in range(miss_row):
            a.loc[len(a)] = ["", "", "", ""]
        a.loc[len(a)+1] = ["*****", f"END TOP O&D {l.upper()}","*****" , "*****"]
        a.loc[len(a) + 2] = ["", "", "", ""]
        l_top.append(a)

    frame = pd.concat(l_top, axis=0, join="inner")
    frame.to_csv(f"csv/BR/br_result/{filename}/top_od.csv")

def excel(filename):
    path = f'csv/BR/br_result/{filename}/'
    list,sheet = [],[]

    for (root, dirs, file) in os.walk(path):
        for f in file:
            if '.csv' in f:
                list.append(f)
    wb = openpyxl.Workbook()
    Sheet_name = wb.sheetnames
    wb.save(filename=f'csv/BR/br_result/{filename}/{filename}.xlsx')
    writer = pd.ExcelWriter(f'csv/BR/br_result/{filename}/{filename}.xlsx', engine='xlsxwriter')
    for i in range(len(list)):
        df = pd.read_csv(f'csv/BR/br_result/{filename}/{list[i]}')
        out_file = list[i].replace('output_custom-extract-','').replace('.csv','').replace('output_','')
        sheet.append(out_file)
        df = df.drop(columns=["Unnamed: 0"])
        df.to_excel(writer, sheet_name = f'{out_file}')
    writer.close()

    wb = load_workbook(f'csv/BR/br_result/{filename}/{filename}.xlsx')

    for i in sheet:
        if "calcul" in i:
            ws1 = wb[i]
            row_total = (ws1.max_row)
            ws1[f'B{row_total}'] = "**Total**"
            ws1[f'K{row_total}'] = ""
            ws1[f'I{row_total}'] = ""
            ws1[f'J{row_total}'] = ""

    for g in range(len(sheet)):
        ws1 = wb[f'{sheet[g]}']
        column = 1
        while column < 59:
            i = get_column_letter(column)
            ws1.column_dimensions[i].width = 20
            column += 1
        ws1.column_dimensions["B"].width = 20
        ws1['A1'] = sheet[g]
        ft = Font(color="ffffff", bold=True,name='Calibri')
        ws1[f'A1'].font = ft
        ws1[f'A1'].fill = PatternFill("solid", start_color="604fd7")
        wb.save(f'csv/BR/br_result/{filename}/{filename}.xlsx')

def write(filename):
    df_resa = pd.read_csv(f"csv/BR/br_result/{filename}/calc_resa.csv")
    df_other = pd.read_csv(f"csv/BR/br_result/{filename}/calc_other.csv")

    TMPLFILE = 'QBR 2022'
    SCOPES = ['https://www.googleapis.com/auth/drive',
              'https://www.googleapis.com/auth/presentations',
              'https://www.googleapis.com/auth/spreadsheets',
             ]

    store = file.Storage('storage.json')
    creds = store.get()

    if not creds or creds.invalid:
        flow = client.flow_from_clientsecrets("creds/credentials.json",SCOPES)
        creds = tools.run_flow(flow, store)

    HTTP = creds.authorize(Http())
    DRIVE = discovery.build('drive', "v3", http=HTTP)
    SHEETS = discovery.build('sheets', "v4", http=HTTP)
    SLIDES = discovery.build('slides', "v1", http=HTTP)

    rsp = DRIVE.files().list (q="name='%s'" % TMPLFILE).execute() ['files'] [0]
    DATA = {'name': f'QR BR {filename}'}
    print('** Copying template %r as %r' % (rsp['name'], DATA['name']))
    DECK_ID = DRIVE.files().copy (body=DATA, fileId=rsp['id']).execute() ['id']

    print('** Replacing placeholder text')
    df_resa = df_resa.loc[df_resa['total_depenses']>0].reset_index()
    sum_reel_resa = df_resa.loc[(df_resa['Type'] == "car")|(df_resa['Type'] == "train")|(df_resa['Type'] == "hotel")|(df_resa['Type'] == "flight")]
    SUM_LEGS= (sum_reel_resa['nbre_trajet'].sum())
    sum_NEW_BILLED= sum_reel_resa['total_depenses'].sum()
    print(sum_NEW_BILLED,SUM_LEGS)

    l_type = [("train", 'T'), ("flight", 'F'), ("hotel", 'H'), ('car', 'C'), ('transfer', 'tf')]
    for type_n,t in l_type:

        try:
            NEW_BILLED = df_resa.loc[df_resa['Type'] == type_n,"total_depenses"].values[0]
        except:
            NEW_BILLED = 0
        try:
            NB_LEGS = df_resa.loc[df_resa['Type'] == type_n,"nbre_trajet"].values[0]
        except:
            NB_LEGS = 0
        try:
            TAUX_ADOPT_CA = ((df_resa.loc[df_resa['Type'] == type_n,"ca_online"].values[0] / df_resa.loc[df_resa['Type'] == type_n,"total_depenses"].values[0])*100).round(0)
            print('taux',TAUX_ADOPT_CA)
        except:
            TAUX_ADOPT_CA = 0
        try:
            ANTICIPATION = df_resa.loc[df_resa['Type'] == type_n,"avg anticipation"].values[0].round(0)
        except:
            ANTICIPATION = 0
        if NEW_BILLED != 0:
            try:
                 AVG_CA = (NEW_BILLED/NB_LEGS).round(2)
            except:
                 AVG_CA= 0


        reqs = [
            # TOTAUX CA
            {"replaceAllText": {
                "containsText": {"text": "{{" + f"total_ca_{t}" + "}}"},
                "replaceText": str(int(NEW_BILLED))}
            },

            # TOTAUX LEGS
            {"replaceAllText": {
                "containsText": {"text": "{{" + f'total_leg_{t}' + "}}"},
                "replaceText": str(int(NB_LEGS))}
            },

            # AVG CA
            {"replaceAllText": {
                "containsText": {"text": "{{" + f'avg_{t}' + "}}"},
                "replaceText": str(AVG_CA)}
            },

            # anticip
            {"replaceAllText": {
                "containsText": {"text": "{{" + f'ant_{t}' + "}}"},
                "replaceText": str(ANTICIPATION)}
            },

            # adopt

            {"replaceAllText": {
                "containsText": {"text": "{{" + f'adop_{t}' + "}}"},
                "replaceText": str(int(TAUX_ADOPT_CA))}
            },


            #total_resa
            {"replaceAllText": {
                "containsText": {"text": "{{" + f'legs' + "}}"},
                "replaceText": str(int(SUM_LEGS))}
            },

            #total_dep
            {"replaceAllText": {
                "containsText": {"text": "{{" + f'Total' + "}}"},
                "replaceText": str(int(sum_NEW_BILLED))}
            }

        ]
        SLIDES.presentations().batchUpdate(body={'requests': reqs},
                                           presentationId=DECK_ID, fields='').execute()

    l_type_o = [("fee", 'f'), ("other_card", 'oc'), ("other_visa", 'ov')]
    for type_n, t in l_type_o:
        try:
            total_fee = df_other.loc[df_other['Type'] == type_n,"total_depenses"].values[0]
        except:
            total_fee = 0
        try:
            avg_fe = df_other.loc[df_other['Type'] == type_n,"avg_dep"].values[0]
        except:
            avg_fe = 0
        try:
            nbre_fe = df_other.loc[df_other['Type'] == type_n,"nbre"].values[0]
        except:
            nbre_fe = 0

        reqs = [
            # TOTAUX FEE
            {"replaceAllText": {
                "containsText": {"text": "{{" + f"total_ca_Fe" + "}}"},
                "replaceText": str(int(total_fee))}
            },

            # TOTAUX LEGS FEE
            {"replaceAllText": {
                "containsText": {"text": "{{" + f'total_leg_Fe' + "}}"},
                "replaceText": str(int(nbre_fe))}
            },

            # AVG CA
            {"replaceAllText": {
                "containsText": {"text": "{{" + f'avg_Fe' + "}}"},
                "replaceText": str(avg_fe)}
            }

        ]
        SLIDES.presentations().batchUpdate(body={'requests': reqs},
                                           presentationId=DECK_ID, fields='').execute()

def create_new_sheet(filename):
    import os.path

    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    import google.auth
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError

    #Connect to gsheet & Drive

    SCOPES = ['https://www.googleapis.com/auth/spreadsheets']

    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'creds/credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        # Save the credentials for the next run
        with open('token.json', 'w') as token:
            token.write(creds.to_json())

    # Create new spreadsheet
    try:
        service = build('sheets', 'v4', credentials=creds)
        spreadsheet = {
            'properties': {
                'title': f"data {filename}"
            }
        }
        spreadsheet = service.spreadsheets().create(body=spreadsheet,
                                                    fields='spreadsheetId') \
            .execute()
        print(f"Spreadsheet ID: {(spreadsheet.get('spreadsheetId'))}")
        id_sp = (spreadsheet.get('spreadsheetId'))
        return spreadsheet.get('spreadsheetId')
    except HttpError as error:
        print(f"An error occurred: {error}")
        return error

def update_sheet(filename):
    import gspread
    from gspread_pandas import Spread

    df = pd.read_csv(f'csv/BR/br_result/{filename}/calcul.csv')
    # print(df)
    gc = gspread.oauth()

    gc.create(f'data {filename}')

    pd_s = Spread(f'data {filename}')
    pd_s.df_to_sheet(df, sheet='Calcul Global', start='A1',freeze_index=True,freeze_headers=True,replace=True)

    sh = gc.open(f'data {filename}')
    worksheet = sh.worksheet("Calcul Global")
    values_list = worksheet.col_values(4)
    for i in values_list:
        if "." in i:
            i = float(i)
            print(i)
    print(values_list)
    # # just gspread for update a cell
    # s_global = client.open('stats SDR 22-23').worksheet('Ringover-Global')
    # s_indiv = client.open('stats SDR 22-23').worksheet('Ringover-Indiv')
    # s_global.update_cell(1,1,f'Updated : {today}')
    # s_indiv.update_cell(1,1,f'Updated : {today}')

def top_hotel_paris():
    datein = "01/01/2022"

    datein = datetime.strptime(datein, '%d/%m/%Y')

    search_expr = re.compile(f".*mama.*", re.I)
    cursor = col_it.find({
        'type': "hotel",
        # '$or':[{"detail.hotelCity" : 'Paris'}, {"detail.hotelAddress" :  ],
        'detail.datein.utc': {'$gte' : datein},
        # "detail.hotelTitle" : {'$regex': search_expr},
        'status' : "confirmed",
        # "detail.hotelRating" : {'$gte': 3},
        "society._id" : ObjectId("5e201fb425fc80001744113e")
    })

    l_hotelTitle,l_hotelAddress,l_hotelPhone,l_hotelRating,l_rooms,l_travelers,l_bookingref,l_supplierbookingref,l_datein,l_hotelCity = [],[],[],[],[],[],[],[],[],[]
    for rep in cursor:
        try:
            hotelTitle = (rep['detail']['hotelTitle'])
        except:
            hotelTitle  = "nc"
        try:
            hotelAddress = rep['detail']['hotelAddress']
        except:
            hotelAddress = "nc"
        try:
            hotelPhone = (rep['detail']['hotelPhone'])
        except:
            hotelPhone = "nc"
        try:
            hotelRating = (rep['detail']['hotelRating'])

        except:
            hotelRating = "nc"
        try:
            hotelCity = (rep['detail']['hotelCity'])

        except:
            hotelCity = "nc"

        try:
            rooms = len(rep['detail']['rooms'])

        except:
            rooms = "nc"
        try:
            travelers = len(rep['travelers'])
            print(travelers)
        except:
            travelers = "nc"
        try:
            bookingref = (rep['detail']['bookingref'])
            print(bookingref)
        except:
            bookingref = "nc"
        try:
            supplierbookingref = (rep['detail']['supplierbookingref'])
            print(supplierbookingref)
        except:
            supplierbookingref = "nc"
        try:
            datein = (rep['detail']['datein']['date'])
            print(datein)
        except:
            datein = "nc"
        l_datein.append(datein)
        l_bookingref.append(bookingref)
        l_supplierbookingref.append(supplierbookingref)
        l_hotelTitle.append(hotelTitle)
        l_hotelAddress.append(hotelAddress)
        l_hotelPhone.append(hotelPhone)
        l_hotelRating.append(hotelRating)
        l_rooms.append(rooms)
        l_travelers.append(travelers)
        l_hotelCity.append(hotelCity)
    df_class = pd.DataFrame(list(zip(l_bookingref,l_supplierbookingref,l_datein,l_hotelTitle,l_hotelAddress,l_hotelCity,l_hotelPhone,l_hotelRating,l_rooms,l_travelers)), columns=['bookingref','supplierbookingref','datein','hotelTitle','hotelAddress','hotelCity','hotelPhone','hotelRating','rooms','travelers'])
    df_class.to_excel('csv/hotel_acton_2022.xlsx')